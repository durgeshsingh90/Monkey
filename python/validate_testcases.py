import pandas as pd
import re
import requests
from collections import defaultdict
import json
import logging

# Configure pandas to print everything
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', None)

ROUTE_TO_API_MAP = {
    "fromiso": "http://localhost:8000/splunkparser/parse/",
    "toiso": "http://localhost:8000/splunkparser/parse/",
    "default": "http://localhost:8000/genericparser/parse_fallback/",  # <-- handles the rest
}

# Custom logging handler to handle Unicode characters
class UnicodeStreamHandler(logging.StreamHandler):
    def emit(self, record):
        try:
            msg = self.format(record)
            stream = self.stream
            if not hasattr(stream, 'encoding') or stream.encoding == 'ANSI_X3.4-1968':
                stream.write(msg.encode('utf-8', 'ignore').decode('utf-8') + self.terminator)
            else:
                stream.write(msg + self.terminator)
            self.flush()
        except Exception:
            self.handleError(record)

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[logging.FileHandler("debug.log", mode='w', encoding='utf-8'),
                              UnicodeStreamHandler()])

# ISO8583 HEADER DETECTION

# Regex pattern to match DE or BM fields in various formats
DE_PATTERN = re.compile(r'\b(?:DE|BM)\s*0*\d+(?:\.\d+)?(?:\s*\([^)]+\))?')

# Function to validate header based on DE pattern
def is_valid_header(row):
    count = 0
    for cell in row:
        if pd.isna(cell):
            continue
        matches = DE_PATTERN.findall(str(cell))
        count += len(matches)
    return count > 5

# Function to read and print Excel data, identifying headers using DE pattern
def read_and_print_excel(file_path):
    all_sheets = pd.read_excel(file_path, sheet_name=None, header=None)
    for sheet_name, df in all_sheets.items():
        header_row_index = None
        for i, row in df.iterrows():
            if is_valid_header(row):
                header_row_index = i
                break

        if header_row_index is not None:
            df.columns = df.iloc[header_row_index]
            df = df[(header_row_index + 1):].reset_index(drop=True)
        else:
            logging.warning("No valid header with >5 ISO8583 DEs found in sheet '%s'.", sheet_name)

# === LOG PARSER + API SENDER ===

TIMESTAMP_PATTERN = re.compile(r'^\d{2}\.\d{2}\.\d{2} \d{2}:\d{2}:\d{2}\.\d{3}')
ROUTE_PATTERN = re.compile(r'\[\s*([A-Za-z]+:\d+)\s*\]')
MSGID_PATTERN = re.compile(r'message id\[(.*?)\]', re.IGNORECASE)

# Function to extract route and message ID from a block of text
def extract_route_and_message_id(block):
    route = None
    message_id = None

    logging.debug("Extracting route and message ID...")
    for line in block.split('\n'):
        line = line.strip()
        logging.debug("Analyzing line: '%s'", line)  # Debug log to see the current line being analyzed
        
        # Extract route if not yet found
        if not route:
            route_match = ROUTE_PATTERN.search(line)
            if route_match:
                route = route_match.group(1).strip()
                logging.debug("Found route: %s", route)  # Debug log for route detection

        # Extract message_id if not yet found
        if not message_id:
            msgid_match = MSGID_PATTERN.search(line)
            if msgid_match:
                message_id = msgid_match.group(1).strip()
                logging.debug("Found message ID: %s", message_id)  # Debug log for message ID detection

        # If both route and message ID are found, no need to continue searching
        if route and message_id:
            break

    logging.debug("Finished route extraction: %s, message ID extraction: %s", route, message_id)
    return route, message_id

# Function to split the log file into blocks based on timestamps
def split_log_blocks(log_file_path):
    with open(log_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    blocks = []
    current_block = []
    for line in lines:
        timestamp_match = TIMESTAMP_PATTERN.match(line)

        if timestamp_match and current_block:
            blocks.append("".join(current_block))
            current_block = []

        current_block.append(line)

    if current_block:
        blocks.append("".join(current_block))

    logging.info("Split into %d blocks", len(blocks))
    return [{"block": block} for block in blocks]

all_responses = []

# Function to send a log block to the API
def send_to_api(log_block, block_index, route=None, message_id=None, url=None, session=None):
    session = session or requests  # fallback in case session is not passed
    try:
        logging.info("\n================== 📨 Sending Block %d ==================", block_index)
        logging.info("📌 Route: %s", route)
        logging.info("📌 Message ID: %s", message_id)
        logging.debug("📜 Block Content:\n%s", log_block)

        response = session.post(url, json={'log_data': log_block})
        response_json = response.json()

        response_json['block_index'] = block_index
        response_json['route'] = route
        response_json['message_id'] = message_id

        all_responses.append(response_json)

        logging.info("\n================== 📩 Response for Block %d ==================", block_index)
        logging.debug(json.dumps(response_json, indent=2))
        logging.info("====================================================\n")

    except Exception as e:
        logging.error("Error sending Block %d: %s", block_index, e)

# Function to process the log file, split into blocks, and send to the API
def process_log_file(log_file_path):
    logging.info("\n📄 Reading Log File and Sending Blocks:")
    blocks = split_log_blocks(log_file_path)
    logging.info("📦 Found %d message blocks.\n", len(blocks))

    session = requests.Session()  # ✅ persistent connection

    for i, block_data in enumerate(blocks, start=1):
        block_content = block_data['block']
        route, message_id = extract_route_and_message_id(block_content)

        if route:
            route_key = route.lower().split(":")[0]
            api_url = ROUTE_TO_API_MAP.get(route_key, ROUTE_TO_API_MAP["default"])
            logging.info("🚀 Sending block %d [Route: %s] to %s", i, route, api_url)
            send_to_api(block_content, i, route, message_id, url=api_url, session=session)
        else:
            logging.info("⏭️ Skipping block %d — No route detected.", i)

    session.close()  # ✅ close when done

def is_hex(s):
    try:
        int(s, 16)
        return True
    except ValueError:
        return False

# Function to convert HEX to ASCII, if possible
def hex_to_ascii(value):
    try:
        return bytes.fromhex(value).decode('ascii')
    except:
        return value  # fallback if it's not valid hex or not decodable

def ascii_to_hex(s):
    return ''.join(format(ord(c), '02x') for c in s)

def normalize_rrn(rrn, existing_rrns):
    if not isinstance(rrn, str):
        return rrn

    # Case 1: Already present as-is (normal RRN)
    if rrn in existing_rrns:
        return rrn

    # Case 2: Looks like hex → decode it
    if is_hex(rrn):
        try:
            decoded = bytes.fromhex(rrn).decode('ascii')
            if decoded.isprintable() and decoded.isdigit():
                return decoded
        except:
            pass

    # Case 3: Convert normal RRN to hex and check reverse match
    hex_version = ascii_to_hex(rrn)
    for existing in existing_rrns:
        if ascii_to_hex(existing) == rrn:
            return existing

    return rrn  # fallback

# Function to group responses by RRN or STAN
def group_responses_by_rrn(responses):
    grouped = defaultdict(list)
    rrn_to_stan = {}
    existing_rrns = set()

    for response in responses:
        data_elements = response.get('result', {}).get('data_elements', {})
        rrn = data_elements.get('DE037')
        stan = data_elements.get('DE011')

        if rrn:
            normalized_rrn = normalize_rrn(rrn, existing_rrns)
            group_key = f"RRN_{normalized_rrn}"
            grouped[group_key].append(response)
            existing_rrns.add(normalized_rrn)

            if stan:
                rrn_to_stan[stan] = group_key

        elif stan:
            matched_group = rrn_to_stan.get(stan)
            if matched_group:
                grouped[matched_group].append(response)
            else:
                grouped[f"STAN_{stan}"].append(response)

        else:
            grouped['UNKNOWN'].append(response)

    # ✅ Log summary
    logging.info("\n📊 Grouping Summary:")
    for group_key, entries in grouped.items():
        logging.info("• %s → %d block(s)", group_key, len(entries))

    return dict(grouped)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Fills for match/mismatch
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# === Get FROMISO block from a group
def get_fromiso_block(group):
    for block in group:
        route = block.get("route", "").lower()
        if route.startswith("fromiso"):
            return block
    return group[0]

# === Load Excel and find matching row based on RRN
def load_excel_and_find_row(rrn, excel_path):
    df = pd.read_excel(excel_path, sheet_name=0)
    df = df.dropna(how="all")
    for i, row in df.iterrows():
        for col in df.columns:
            if str(row[col]).strip() == str(rrn):
                return df, i
    return df, None

# === Apply green/red coloring to matched/mismatched cells
def apply_color_coding_with_status(df, row_index, parsed_data, excel_path, status_map):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = load_workbook(excel_path)
    ws = wb.active

    header_row = 1
    status_col_index = len(df.columns) + 1
    ws.cell(row=header_row, column=status_col_index).value = "ValidationStatus"

    all_match = True
    for col_index, col_name in enumerate(df.columns, start=1):
        expected = str(df.loc[row_index, col_name]).strip()
        key = col_name.split()[0].replace("BM", "DE")
        actual = str(parsed_data.get(key, "")).strip()

        if expected == actual or expected.lower() == "client-defined":
            ws.cell(row=row_index + 2, column=col_index).fill = green_fill
        else:
            all_match = False
            ws.cell(row=row_index + 2, column=col_index).fill = red_fill

    if all_match:
        status = "Passed"
    else:
        status = "Failed"

    ws.cell(row=row_index + 2, column=status_col_index).value = status
    status_map[row_index] = status

    wb.save("validated_output.xlsx")

# === Validate each group by matching FromISO data with Excel test case
def validate_groups_against_excel(grouped_data, excel_path):
    status_map = {}
    df = pd.read_excel(excel_path, sheet_name=0)
    df = df.dropna(how="all")

    matched_rows = set()

    for rrn_group_key, group_blocks in grouped_data.items():
        if not rrn_group_key.startswith("RRN_"):
            continue

        rrn = rrn_group_key.replace("RRN_", "")
        found = False

        for i, row in df.iterrows():
            for col in df.columns:
                if str(row[col]).strip() == str(rrn):
                    fromiso_block = get_fromiso_block(group_blocks)
                    parsed_data = fromiso_block.get("result", {}).get("data_elements", {})
                    apply_color_coding_with_status(df, i, parsed_data, excel_path, status_map)
                    matched_rows.add(i)
                    found = True
                    break
            if found:
                break

        if not found:
            logging.warning("⚠️ No matching row found in Excel for RRN: %s", rrn)

    # Mark skipped rows
    wb = load_workbook(excel_path)
    ws = wb.active
    status_col_index = len(df.columns) + 1
    ws.cell(row=1, column=status_col_index).value = "ValidationStatus"
    for i in range(len(df)):
        if i not in matched_rows:
            ws.cell(row=i + 2, column=status_col_index).value = "Skipped"

    wb.save("validated_output.xlsx")

# === Final call added to your existing __main__ block:
if __name__ == "__main__":
    excel_file_path = r"D:\Projects\VSCode\MangoData\ISO8583_eCommerce_TestCases (1).xlsx"
    log_file_path = r"D:\Projects\VSCode\MangoData\splunk_log.txt"
    read_and_print_excel(excel_file_path)

    all_responses = []
    process_log_file(log_file_path)
    grouped_data = group_responses_by_rrn(all_responses)

    with open("grouped_by_rrn.json", "w") as f:
        json.dump(grouped_data, f, indent=2)

    logging.info("✅ Grouped data saved to grouped_by_rrn.json")

    # ✅ Add this at the end of your script
    validate_groups_against_excel(grouped_data, excel_file_path)
    logging.info("✅ Validation complete. Check 'validated_output.xlsx' for colored results.")
